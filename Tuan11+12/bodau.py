import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import openpyxl
import os
from unidecode import unidecode
from datetime import datetime  # 🕒 thêm timestamp

# Cấu hình đường dẫn Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Đường dẫn tệp và thư mục
input_pdf = r"C:/Users/Admin/OneDrive/Documents/GitHub/GR1/output/pages/page_8.pdf"
output_dir = r"C:/Users/Admin/OneDrive/Documents/GitHub/GR1/output_images"
result_dir = r"C:/Users/Admin/OneDrive/Documents/GitHub/GR1/output"  # 📁 thư mục chứa Excel

# Đảm bảo thư mục tồn tại
os.makedirs(output_dir, exist_ok=True)
os.makedirs(result_dir, exist_ok=True)

# Tạo tên file Excel mới dựa theo thời gian
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_output = os.path.join(result_dir, f"result_{timestamp}.xlsx")

# Bước 1: Chuyển PDF thành ảnh
images = convert_from_path(input_pdf, poppler_path=r"D:\Downloads\poppler-24.08.0\Library\bin")  # ⚠️ sửa nếu cần

# Bước 2: Tạo workbook Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Extracted Text"

# Ghi tiêu đề cột
ws.cell(row=1, column=1, value="Văn bản gốc")
ws.cell(row=1, column=2, value="Văn bản không dấu")
row_counter = 2

# Bước 3: OCR và ghi kết quả
for i, img in enumerate(images):
    img_path = os.path.join(output_dir, f"page_{i+1}.png")
    img.save(img_path, "PNG")

    text = pytesseract.image_to_string(Image.open(img_path), lang='vie')  # OCR tiếng Việt

    for line in text.split("\n"):
        line = line.strip()
        if line:
            line_khong_dau = unidecode(line)
            ws.cell(row=row_counter, column=1, value=line)
            ws.cell(row=row_counter, column=2, value=line_khong_dau)
            row_counter += 1

# Bước 4: Lưu file Excel
wb.save(excel_output)
print(f"✅ Đã tạo file Excel mới: {excel_output}")
